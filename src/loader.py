import os
from openpyxl import load_workbook
from config import OUTPUT_DIR, TOP_NUMBER

def load(df, country_name, OUTPUT_DIR, TOP_NUMBER):
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        file_name = country_name.replace(" ", "_")
        file_path = f"{OUTPUT_DIR}{file_name}_top{TOP_NUMBER}.xlsx"
        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '#,##0.00 "USD"'
        wb.save(file_path)

        print(f"✅ Saved: {file_path}")
    except PermissionError:
        print(f"❌ File is open in Excel, close it first: {file_path}")
        raise
    except Exception as e:
        print(f"❌ Error saving {country_name}: {e}")
        raise
